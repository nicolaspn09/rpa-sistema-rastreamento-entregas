import os
import openpyxl

# Diretórios das pastas
pasta_comparacao = r'C:\rpa\CD\Acompanhamento de Entregas\Arquivos Comparação'
pasta_download = r'C:\rpa\CD\Acompanhamento de Entregas\Download Site Comprovei'

# Lista de transportadores
transportadores = ["Rondolog", "Translog", "BSB"]

# Dicionário para armazenar as informações das notas fiscais
notas_fiscais = {}

# Loop pelos transportadores
for transportador in transportadores:
    contador = 0
    arquivo_comparacao = os.path.join(pasta_comparacao, f"{transportador}.xlsx")
    arquivo_download = os.path.join(pasta_download, f"{transportador}.xlsx")

    if os.path.exists(arquivo_comparacao) and os.path.exists(arquivo_download):
        # Carregue os arquivos Excel
        wb_comparacao = openpyxl.load_workbook(arquivo_comparacao)
        wb_download = openpyxl.load_workbook(arquivo_download)

        # Acesse as planilhas (ajuste os nomes das planilhas conforme necessário)
        planilha_comparacao = wb_comparacao.active
        planilha_download = wb_download.active

        quantidade_linhas_planilha_download = planilha_download.max_row

        # Percorra as linhas na planilha de comparação
        for linha in planilha_comparacao.iter_rows(min_row=2, values_only=True):
            nota_fiscal = linha[7]  

            # Percorra as linhas na planilha de download
            for linha in planilha_download.iter_rows(min_row=2, values_only=True):
                contador = contador + 1
                nota_fiscal_comprovei = linha[1]
                situacao_comprovei = linha[17]

                if nota_fiscal == nota_fiscal_comprovei:
                    if situacao_comprovei == 'Chegou' or situacao_comprovei == 'Entregue':
                        print('Entregue')
                        break
                    else:
                        print(situacao_comprovei)
                        break
                else:
                    if contador == quantidade_linhas_planilha_download:
                        print('Não encontrado')
                        break


            # Feche os arquivos Excel
            wb_comparacao.close()
            wb_download.close()

# Atualize as informações na planilha "Arquivos Comparação"
for transportador in transportadores:
    arquivo_comparacao = os.path.join(pasta_comparacao, f"{transportador}.xlsx")

    if os.path.exists(arquivo_comparacao):
        wb_comparacao = openpyxl.load_workbook(arquivo_comparacao)
        planilha_comparacao = wb_comparacao.active

        for row in planilha_comparacao.iter_rows(min_row=2, values_only=True):
            nota_fiscal = row[7]

            if nota_fiscal in notas_fiscais:
                situacao = notas_fiscais[nota_fiscal]["situacao"]
                planilha_comparacao.cell(row=row[0].row, column=18, value=situacao)

        wb_comparacao.save(arquivo_comparacao)
        wb_comparacao.close()