import os
import openpyxl

# Diretórios das pastas
pasta_comparacao = r'C:\rpa\CD\Acompanhamento de Entregas\Arquivos Comparação'
pasta_download = r'C:\rpa\CD\Acompanhamento de Entregas\Download Site Comprovei'

# Lista de transportadores
transportadores = ["Rondolog", "Translog", "BSB"]

# Loop pelos transportadores
for transportador in transportadores:
    contador = 0
    arquivo_comparacao = os.path.join(pasta_comparacao, f"{transportador}.xlsx")
    arquivo_download = os.path.join(pasta_download, f"{transportador}.xlsx")

    if os.path.exists(arquivo_comparacao) and os.path.exists(arquivo_download):
        # Carregue os arquivos Excel
        wb_comparacao = openpyxl.load_workbook(arquivo_comparacao)
        wb_download = openpyxl.load_workbook(arquivo_download)

        # Acesse as planilhas (ajuste os nomes das planilhas conforme necessário)
        planilha_comparacao = wb_comparacao.active
        planilha_download = wb_download.active

        quantidade_linhas_planilha_download = planilha_download.max_row

        # Percorra as linhas na planilha de comparação
        for linha_comparacao in planilha_comparacao.iter_rows(min_row=2, values_only=True):
            nota_fiscal = linha_comparacao[7]
            nota_fiscal = nota_fiscal.strip()
            linha_comparacao_informar = list(linha_comparacao)

            # Percorra as linhas na planilha de download
            for linha_download in planilha_download.iter_rows(min_row=2, values_only=True):
                contador += 1
                nota_fiscal_comprovei = linha_download[1]
                nota_fiscal_comprovei = str(nota_fiscal_comprovei).strip()
                situacao_comprovei = linha_download[17]
                situacao_comprovei = situacao_comprovei.strip()

                if nota_fiscal == nota_fiscal_comprovei:
                    if situacao_comprovei == 'Chegou' or situacao_comprovei == 'Entregue':
                        # Atualize a coluna 18 na planilha de comparação com a situação
                        linha_comparacao_informar[18] = situacao_comprovei
                        break
                    else:
                        linha_comparacao_informar[18] = situacao_comprovei
                        break
                else:
                    if contador == quantidade_linhas_planilha_download:
                        situacao_comprovei = 'Não encontrado'
                        linha_comparacao_informar[18] = situacao_comprovei
                        break

        # Volte ao início da planilha de download para a próxima iteração da planilha de comparação
        planilha_download = wb_download.active

        # Feche os arquivos Excel
        wb_comparacao.save(arquivo_comparacao)
        wb_comparacao.close()
        wb_download.close()